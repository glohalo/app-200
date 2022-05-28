# -*- coding: utf-8 -*-
"""
Created on Mon Jan 17 14:20:17 2022

@author: Administrador
"""

#creating a xlm file from .xlsm
#First, I need to import modules
#Openpyxl helps to interacting with excel files
#yattag, allows me to convert to HTML and xlm
from openpyxl import load_workbook
from yattag import Doc, indent


#loading the excel file
wb = load_workbook("demo_database.xlsx")
#getting an object of active sheet 1
wb = wb.worksheets[0]


#Returning returns a triplet
#The tagtext() method is a helper method that retunrns a triplet composed of:
    #- the Doc instance itself
    #- The tag method of the Doc instance
    #-The text method of the Doc instance
    
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

# Appends the String to document
doc.asis(xml_header)
doc.asis(xml_schema)

with tag('People'):
    for row in wb.iter_rows(min_row=2, max_row=10, min_col=1, max_col=6):
        row = [cell.value for cell in row]
        with tag("Person"):
            with tag("First_Name"):
                text(row[0])
            with tag("Last_Name"):
                text(row[1])
            with tag("Gender"):
                text(row[2])
            with tag("Country"):
                text(row[3])
            with tag("Age"):
                text(row[4])
            with tag("Date"):
                text(row[5])
                
result = indent(
    doc.getvalue(),
    indentation='   ',
    indent_text=True
)
  
with open("output.xml", "w") as f:
    f.write(result)
    
    
    
##Reading xlm file
#did this with the proof file
 
#finding tags
##extracting tags
from bs4 import BeautifulSoup

#reading the data inside the xml file 
#to a variable under the name data

with open('DOCUMENTO_CONTABLE.txt', 'r') as f:
    data = f.read()
    
#passing the stored data inside the beautifulSoap parser, storing the returned
#object

Bs_data = BeautifulSoup(data, "xml")

#Finding all instances of tag, "Unique"
b_unique = Bs_data.find_all('linea')
 
print(b_unique)

# Using find() to extract attributes
# of the first instance of the tag
b_name = Bs_data.find('Datos')

#print(b_name)

#Extracting the data stored in a specific attribute of the any tag
value = b_name.get('00000023501')
 
print(value)
